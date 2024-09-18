import tkinter as tk
from tkinter import scrolledtext, messagebox
from tkinter import PhotoImage
from openpyxl import load_workbook
from prettytable import PrettyTable
import datetime
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

# Carregar a planilha Excel
tabela = load_workbook("GESTAO_DE_EXAMES_PERIODICOS.xlsx", data_only=True)
aba_ativa = tabela.active

# Função para formatar datas no padrão dd-mm-aaaa
def formatar_data(valor):
    if isinstance(valor, datetime.datetime):
        return valor.strftime('%d-%m-%Y')
    return valor

# Função para calcular o tempo restante entre duas datas
def calcular_tempo_restante(data_inicio, data_fim):
    if isinstance(data_inicio, datetime.datetime) and isinstance(data_fim, datetime.datetime):
        return (data_fim - data_inicio).days
    return ""

# Função para filtrar as linhas onde o valor da coluna I é <= 60
def filtrar_linhas():
    linhas_filtradas = []
    for linha in aba_ativa.iter_rows(min_row=5, max_col=9, values_only=True):
        if linha[8] and isinstance(linha[8], int) and linha[8] <= 60:  # Verifica se o valor da coluna I (índice 8) é <= 60
            linha_formatada = [formatar_data(celula) for celula in linha[:9]]
            linhas_filtradas.append(linha_formatada)
    return linhas_filtradas

# Função para enviar o e-mail com as linhas filtradas
def enviar_email(destinatario):
    linhas_filtradas = filtrar_linhas()
    if not linhas_filtradas:
        messagebox.showinfo("Informação", "Nenhuma linha encontrada com valor <= 60 na coluna I.")
        return

    corpo_email = "Segue abaixo a lista de colaboradores com tempo restante igual ou menor a 60 dias:\n\n"
    
    for linha in linhas_filtradas:
        corpo_email += " | ".join(str(campo) if campo is not None else "" for campo in linha) + "\n"

    # Configurações de e-mail
    remetente = "seuemail@gmail.com"  # Seu e-mail
    senha = "suasenha"  # Sua senha de e-mail

    mensagem = MIMEMultipart()
    mensagem['From'] = remetente
    mensagem['To'] = destinatario
    mensagem['Subject'] = "Colaboradores com tempo restante <= 60 dias"

    # Adicionar o corpo ao e-mail
    mensagem.attach(MIMEText(corpo_email, 'plain'))

    try:
        # Conectar ao servidor SMTP do Gmail
        servidor = smtplib.SMTP('smtp.gmail.com', 587)
        servidor.starttls()
        servidor.login(remetente, senha)

        # Enviar o e-mail
        servidor.send_message(mensagem)
        servidor.quit()

        messagebox.showinfo("Sucesso", "E-mail enviado com sucesso!")
    except Exception as e:
        messagebox.showerror("Erro", f"Falha ao enviar o e-mail: {str(e)}")

# Função para atualizar dados da tabela
def atualizar_dados():
    global linhas_formatadas, colunas_validas
    linhas_formatadas = []
    for linha in aba_ativa.iter_rows(min_row=5, max_col=9, values_only=True):  # Limitando as colunas até a 9ª
        linha_formatada = [formatar_data(celula) for celula in linha]

        # Cálculo do tempo restante (5ª e 6ª colunas)
        if linha[4] and linha[5]:  # Verifica se ambas as datas estão preenchidas
            data_inicio = linha[4]
            data_fim = linha[5]
            tempo_restante = calcular_tempo_restante(data_inicio, data_fim)
            linha_formatada.append(str(tempo_restante))  # Adiciona o tempo restante à última coluna
        else:
            linha_formatada.append("")  # Adiciona uma célula vazia se as datas não existirem

        if any(celula is not None for celula in linha_formatada):
            linhas_formatadas.append(linha_formatada)

    # Ordenar as linhas formatadas pela primeira coluna
    linhas_formatadas.sort(key=lambda x: str(x[0]).lower() if x[0] is not None else "")

    colunas_validas = []
    for i in range(len(cabecalhos_unicos)):
        if i < 9:  # Limitar o número de colunas até 9
            if any(linha[i] is not None for linha in linhas_formatadas):
                colunas_validas.append(i)

# Definir cabeçalhos da tabela com base na 4ª linha
cabecalhos = [celula.value if celula.value is not None else f"Coluna_{i}" for i, celula in enumerate(aba_ativa[4][:9], start=1)]
cabecalhos_unicos = []
for i, cabecalho in enumerate(cabecalhos):
    if cabecalho in cabecalhos_unicos:
        cabecalhos_unicos.append(f"{cabecalho}_{i+1}")
    else:
        cabecalhos_unicos.append(cabecalho)

# Função para exibir a tabela na interface tkinter
def exibir_tabela():
    atualizar_dados()
    tabela_sem_colunas_vazias = PrettyTable()
    tabela_sem_colunas_vazias.field_names = [cabecalhos_unicos[i] for i in colunas_validas]
    for linha in linhas_formatadas:
        linha_filtrada = [linha[i] for i in colunas_validas]
        tabela_sem_colunas_vazias.add_row(linha_filtrada)

    janela_exibir = tk.Toplevel()
    janela_exibir.title("Exibição da Tabela")

    frame = tk.Frame(janela_exibir)
    frame.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

    scroll_x = tk.Scrollbar(frame, orient=tk.HORIZONTAL)
    scroll_x.pack(side=tk.BOTTOM, fill=tk.X)
    scroll_y = tk.Scrollbar(frame)
    scroll_y.pack(side=tk.RIGHT, fill=tk.Y)

    txt_tabela = tk.Text(frame, wrap=tk.NONE, xscrollcommand=scroll_x.set, yscrollcommand=scroll_y.set, width=160, height=40)
    txt_tabela.pack(fill=tk.BOTH, expand=True)
    scroll_x.config(command=txt_tabela.xview)
    scroll_y.config(command=txt_tabela.yview)

    cor_laranja_claro = "#FFCC99"
    cor_padrao = "#FFFFFF"

    for i, linha in enumerate(str(tabela_sem_colunas_vazias).split("\n")):
        if i % 2 == 0:
            txt_tabela.insert(tk.END, linha + "\n", ("laranja",))
        else:
            txt_tabela.insert(tk.END, linha + "\n", ("branco",))

    txt_tabela.tag_configure("laranja", background=cor_laranja_claro)
    txt_tabela.tag_configure("branco", background=cor_padrao)
    txt_tabela.config(state=tk.DISABLED)

# Função para criar a tela de envio de e-mail
def tela_envio_email():
    janela_email = tk.Toplevel()
    janela_email.title("Enviar e-mail")

    label_email = tk.Label(janela_email, text="Enviar para:")
    label_email.grid(row=0, column=0)
    entrada_email = tk.Entry(janela_email)
    entrada_email.grid(row=0, column=1)

    def enviar():
        destinatario = entrada_email.get()
        if destinatario:
            enviar_email(destinatario)
            janela_email.destroy()
        else:
            messagebox.showerror("Erro", "Por favor, insira um endereço de e-mail válido.")

    botao_enviar = tk.Button(janela_email, text="Enviar", command=enviar)
    botao_enviar.grid(row=1, column=1)

# Função para criar uma nova linha
def criar_linha():
    janela_criar = tk.Toplevel()
    janela_criar.title("Acrescentar Colaborador e Seus Dados")

    campos = []

    for i, cabecalho in enumerate(cabecalhos_unicos):
        label = tk.Label(janela_criar, text=cabecalho)
        label.grid(row=i, column=0)
        campo = tk.Entry(janela_criar)
        campo.grid(row=i, column=1)
        campos.append(campo)

    def salvar_linha():
        nova_linha = [campo.get() for campo in campos]
        aba_ativa.append(nova_linha)
        tabela.save("GESTAO_DE_EXAMES_PERIODICOS.xlsx")
        messagebox.showinfo("Sucesso", "Linha criada com sucesso!")
        janela_criar.destroy()

    botao_salvar = tk.Button(janela_criar, text="Salvar", command=salvar_linha)
    botao_salvar.grid(row=len(cabecalhos_unicos), column=1)

# Função para alterar uma linha existente (selecionando pela primeira coluna)
def alterar_linha():
    janela_alterar = tk.Toplevel()
    janela_alterar.title("Alterar Informações")

    label_linha = tk.Label(janela_alterar, text="Nome Completo Colaborador:")
    label_linha.grid(row=0, column=0)
    entrada_valor = tk.Entry(janela_alterar)
    entrada_valor.grid(row=0, column=1)

    campos = []
    for i, cabecalho in enumerate(cabecalhos_unicos):
        label = tk.Label(janela_alterar, text=cabecalho)
        label.grid(row=i+1, column=0)
        campo = tk.Entry(janela_alterar)
        campo.grid(row=i+1, column=1)
        campos.append(campo)

    def salvar_alteracao():
        valor_chave = entrada_valor.get()
        for linha in aba_ativa.iter_rows(min_row=5, max_col=9):
            if str(linha[0].value) == valor_chave:
                for i, campo in enumerate(campos):
                    aba_ativa.cell(row=linha[0].row, column=i+1, value=campo.get())
                tabela.save("GESTAO_DE_EXAMES_PERIODICOS.xlsx")
                messagebox.showinfo("Sucesso", "Linha alterada com sucesso!")
                janela_alterar.destroy()
                return
        messagebox.showerror("Erro", "Valor não encontrado!")

    botao_salvar = tk.Button(janela_alterar, text="Salvar Alterações", command=salvar_alteracao)
    botao_salvar.grid(row=len(cabecalhos_unicos)+1, column=1)

# Função para excluir uma linha existente (selecionando pela primeira coluna)
def excluir_linha():
    janela_excluir = tk.Toplevel()
    janela_excluir.title("Excluir Colaborador")

    label_linha = tk.Label(janela_excluir, text="Nome Completo Colaborador para excluir:")
    label_linha.grid(row=0, column=0)
    entrada_valor = tk.Entry(janela_excluir)
    entrada_valor.grid(row=0, column=1)

    def confirmar_exclusao():
        valor_chave = entrada_valor.get()
        for linha in aba_ativa.iter_rows(min_row=5, max_col=9):
            if str(linha[0].value) == valor_chave:
                aba_ativa.delete_rows(linha[0].row)
                tabela.save("GESTAO_DE_EXAMES_PERIODICOS.xlsx")
                messagebox.showinfo("Sucesso", "Linha excluída com sucesso!")
                janela_excluir.destroy()
                return
        messagebox.showerror("Erro", "Valor não encontrado!")

    botao_excluir = tk.Button(janela_excluir, text="Excluir", command=confirmar_exclusao)
    botao_excluir.grid(row=1, column=1)

# Função para criar a tela inicial com os botões de ação
def tela_inicial():
    janela = tk.Tk()
    janela.title("GESTÃO EXAMES PERIÓDICOS")

    # Carregar a imagem do logotipo 
    logotipo = PhotoImage(file="LogoGM.png")
    label_logo = tk.Label(janela, image=logotipo)
    label_logo.grid(row=0, column=0, rowspan=5, padx=20, pady=20, sticky="w")  # Logotipo à esquerda

    # Criação dos botões e alinhamento à direita
    botao_exibir = tk.Button(janela, text="Exibir Tabela", command=exibir_tabela, width=20)
    botao_exibir.grid(row=0, column=1, padx=10, pady=10, sticky="e")

    botao_criar = tk.Button(janela, text="Acrescentar Colaborador", command=criar_linha, width=20)
    botao_criar.grid(row=1, column=1, padx=10, pady=10, sticky="e")

    botao_alterar = tk.Button(janela, text="Alterar Informações", command=alterar_linha, width=20)
    botao_alterar.grid(row=2, column=1, padx=10, pady=10, sticky="e")

    botao_excluir = tk.Button(janela, text="Excluir Colaborador", command=excluir_linha, width=20)
    botao_excluir.grid(row=3, column=1, padx=10, pady=10, sticky="e")

    botao_enviar_email = tk.Button(janela, text="Enviar Relatório por E-mail", command=tela_envio_email, width=20)
    botao_enviar_email.grid(row=4, column=1, padx=10, pady=10, sticky="e")

    janela.mainloop()

# Iniciar a tela inicial
tela_inicial()